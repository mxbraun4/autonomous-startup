"""Generate collaboration framework Excel from run data."""
import json
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter

wb = Workbook()
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
hfont = Font(bold=True, color="FFFFFF", size=11)
fill_header = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

phase_colors = {
    "FORMING": "B4C7E7",
    "STORMING": "F4B183",
    "NORMING": "A9D18E",
    "PERFORMING": "FFD966",
    "SUSTAINING": "D5A6BD",
}

# ---- Sheet 1: Per-Iteration Classification ----
ws = wb.active
ws.title = "Iteration Classification"

iter_data = [
    (1, "Greenfield Build", "Full Pipeline", "Initial Build", "FORMING", 3, "strategist > developer > reviewer"),
    (2, "Critical Rewrite", "Direct Fix", "Recurring Bug", "STORMING", 1, "developer"),
    (3, "Bug Fix", "Full Pipeline", "Customer Feedback", "STORMING", 3, "strategist > developer > reviewer"),
    (4, "Critical Rewrite", "Direct Fix", "Customer Feedback", "STORMING", 1, "developer"),
    (5, "Bug Fix", "Diagnose + Verify", "Customer Feedback", "STORMING", 2, "strategist > reviewer"),
    (6, "Bug Fix", "Fix + Verify", "Customer Feedback", "STORMING", 2, "reviewer > developer"),
    (7, "Critical Rewrite", "Direct Fix", "Customer Feedback", "STORMING", 1, "developer"),
    (8, "Incremental Improvement", "Full Pipeline", "Recurring Bug", "NORMING", 3, "strategist > developer > reviewer"),
    (9, "Bug Fix", "Full Pipeline", "Recurring Bug", "NORMING", 3, "strategist > developer > reviewer"),
    (10, "Bug Fix", "Full Pipeline", "Recurring Bug", "NORMING", 3, "strategist > developer > reviewer"),
    (11, "Critical Rewrite", "Full Pipeline", "Customer Feedback", "NORMING", 3, "strategist > developer > reviewer"),
    (12, "Critical Rewrite", "Full Pipeline", "Recurring Bug", "NORMING", 4, "strategist > developer > reviewer > developer"),
    (13, "Bug Fix", "Full Pipeline", "Recurring Bug", "NORMING", 3, "strategist > developer > reviewer"),
    (14, "Bug Fix", "Diagnose + Verify", "Customer Feedback", "NORMING", 2, "strategist > reviewer"),
    (15, "Critical Rewrite", "Fix + Verify", "Recurring Bug", "NORMING", 2, "developer > reviewer"),
    (16, "Critical Rewrite", "Full Pipeline", "Customer Feedback", "PERFORMING", 3, "strategist > developer > reviewer"),
    (17, "Critical Rewrite", "Fix + Verify", "Customer Feedback", "PERFORMING", 2, "developer > reviewer"),
    (18, "Bug Fix", "Fix + Verify", "Customer Feedback", "PERFORMING", 2, "developer > reviewer"),
    (19, "Bug Fix", "Fix + Verify", "Recurring Bug", "PERFORMING", 4, "developer > reviewer > developer > reviewer"),
    (20, "Feature Addition", "Fix + Verify", "Team Knowledge", "PERFORMING", 2, "developer > reviewer"),
    (21, "Critical Rewrite", "Full Pipeline", "Customer Feedback", "PERFORMING", 4, "strategist > developer > reviewer > strategist"),
    (22, "Critical Rewrite", "Full Pipeline", "Customer Feedback", "PERFORMING", 3, "strategist > developer > reviewer"),
    (23, "Critical Rewrite", "Full Pipeline", "Customer Feedback", "PERFORMING", 3, "strategist > developer > reviewer"),
    (24, "Bug Fix", "Full Pipeline", "Customer Feedback", "SUSTAINING", 3, "strategist > developer > reviewer"),
    (25, "Bug Fix", "Direct Fix", "Customer Feedback", "SUSTAINING", 1, "developer"),
    (26, "Critical Rewrite", "Fix + Verify", "Customer Feedback", "SUSTAINING", 2, "developer > reviewer"),
    (27, "Bug Fix", "Full Pipeline", "Customer Feedback", "SUSTAINING", 3, "strategist > developer > reviewer"),
    (28, "Bug Fix", "Verification Only", "Customer Feedback", "SUSTAINING", 1, "reviewer"),
    (29, "Bug Fix", "Fix + Verify", "Customer Feedback", "SUSTAINING", 2, "developer > reviewer"),
]

headers = ["Iteration", "Work Type", "Collaboration Pattern", "Trigger", "Development Phase", "Dispatches", "Dispatch Chain"]

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hfont
    c.fill = fill_header
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = border

for ri, row in enumerate(iter_data, 2):
    phase = row[4]
    fill = PatternFill(start_color=phase_colors.get(phase, "FFFFFF"), end_color=phase_colors.get(phase, "FFFFFF"), fill_type="solid")
    for col, val in enumerate(row, 1):
        c = ws.cell(row=ri, column=col, value=val)
        c.alignment = Alignment(horizontal="center" if col < 7 else "left", wrap_text=True)
        c.border = border
        c.fill = fill

for col, w in enumerate([12, 24, 24, 20, 20, 12, 50], 1):
    ws.column_dimensions[get_column_letter(col)].width = w
ws.freeze_panes = "A2"

# ---- Sheet 2: Framework Summary (Tuckman-inspired) ----
ws2 = wb.create_sheet("Framework Summary")

phases_summary = [
    ("FORMING", "1", "1", "Scaffolding",
     "Full pipeline builds from scratch",
     "Strategist designs, Developer builds, Reviewer validates",
     "Initial Build"),
    ("STORMING", "2-7", "6", "Bug Whack-a-Mole",
     "Coordinator tries different dispatch patterns",
     "50% Direct Fix - team searching for rhythm",
     "Customer Feedback (67%)"),
    ("NORMING", "8-15", "8", "Systematic Response",
     "Full Pipeline becomes default pattern (75%)",
     "Strategist diagnoses root cause before Developer acts",
     "Recurring Bug (50%)"),
    ("PERFORMING", "16-23", "8", "Feature + Polish",
     "Mix of Full Pipeline (50%) and Fix+Verify (50%)",
     "Feature additions, UI polish, targeted rewrites",
     "Customer Feedback (75%)"),
    ("SUSTAINING", "24-29", "6", "Maintenance Mode",
     "Lighter patterns: Direct Fix, Verification Only emerge",
     "Fewer dispatches, more verification, product mature",
     "Customer Feedback (83%)"),
]

ph_headers = ["Phase", "Iterations", "Count", "Label", "Collaboration Behavior", "Key Characteristic", "Primary Trigger"]
for col, h in enumerate(ph_headers, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = hfont
    c.fill = fill_header
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = border

for ri, row in enumerate(phases_summary, 2):
    fill = PatternFill(
        start_color=phase_colors.get(row[0], "FFFFFF"),
        end_color=phase_colors.get(row[0], "FFFFFF"),
        fill_type="solid",
    )
    for col, val in enumerate(row, 1):
        c = ws2.cell(row=ri, column=col, value=val)
        c.alignment = Alignment(horizontal="center" if col < 5 else "left", wrap_text=True)
        c.border = border
        c.fill = fill

ws2.row_dimensions[2].height = 30
ws2.row_dimensions[3].height = 30
ws2.row_dimensions[4].height = 30
ws2.row_dimensions[5].height = 30
ws2.row_dimensions[6].height = 30
for col, w in enumerate([16, 12, 8, 24, 45, 45, 28], 1):
    ws2.column_dimensions[get_column_letter(col)].width = w

# ---- Sheet 3: Distribution Charts ----
ws3 = wb.create_sheet("Distributions")

# Work Type
ws3.cell(row=1, column=1, value="Work Type").font = Font(bold=True)
ws3.cell(row=1, column=2, value="Count").font = Font(bold=True)
wt_data = [("Greenfield Build", 1), ("Feature Addition", 1), ("Incremental Improvement", 1), ("Bug Fix", 12), ("Critical Rewrite", 14)]
for i, (wt, c) in enumerate(wt_data, 2):
    ws3.cell(row=i, column=1, value=wt)
    ws3.cell(row=i, column=2, value=c)

pie1 = PieChart()
pie1.title = "Work Type Distribution"
pie1.width = 18
pie1.height = 12
cats = Reference(ws3, min_col=1, min_row=2, max_row=6)
vals = Reference(ws3, min_col=2, min_row=1, max_row=6)
pie1.add_data(vals, titles_from_data=True)
pie1.set_categories(cats)
ws3.add_chart(pie1, "D1")

# Collaboration Pattern
ws3.cell(row=9, column=1, value="Collaboration Pattern").font = Font(bold=True)
ws3.cell(row=9, column=2, value="Count").font = Font(bold=True)
cp_data = [("Full Pipeline", 14), ("Fix + Verify", 8), ("Direct Fix", 4), ("Diagnose + Verify", 2), ("Verification Only", 1)]
for i, (cp, c) in enumerate(cp_data, 10):
    ws3.cell(row=i, column=1, value=cp)
    ws3.cell(row=i, column=2, value=c)

pie2 = PieChart()
pie2.title = "Collaboration Pattern Distribution"
pie2.width = 18
pie2.height = 12
cats2 = Reference(ws3, min_col=1, min_row=10, max_row=14)
vals2 = Reference(ws3, min_col=2, min_row=9, max_row=14)
pie2.add_data(vals2, titles_from_data=True)
pie2.set_categories(cats2)
ws3.add_chart(pie2, "D16")

# Trigger
ws3.cell(row=17, column=1, value="Trigger").font = Font(bold=True)
ws3.cell(row=17, column=2, value="Count").font = Font(bold=True)
tr_data = [("Customer Feedback", 19), ("Recurring Bug", 8), ("Initial Build", 1), ("Team Knowledge", 1)]
for i, (t, c) in enumerate(tr_data, 18):
    ws3.cell(row=i, column=1, value=t)
    ws3.cell(row=i, column=2, value=c)

ws3.column_dimensions["A"].width = 25
ws3.column_dimensions["B"].width = 10

wb.save("collaboration_framework.xlsx")
print("Saved collaboration_framework.xlsx")
